# This program will read all the links from the db
#  Check for valid links
    # get the number of Reviews
    # get the number of ratings
    # get the price
    # update in mongodb
from pymongo import MongoClient
from selenium import webdriver
from selenium import *
import re,datetime,xlrd,time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import *

client = MongoClient('127.0.0.1',27017)
db = client['Flipkart']
col = db['flipkart']
# col2 = db['result']
dt = datetime.datetime.now().date()
driver = webdriver.Chrome()
wb = Workbook()
ws = wb.create_sheet('Sheet1',1)
fn = 'temp1.xlsx'
# rev = str(dt)+'review'
# rat = str(dt)+'rating'
# pri = str(dt)+'price'
ws.cell(row = 1, column=1).value = "id"
ws.cell(row = 1, column=2).value = "brand"
ws.cell(row = 1, column=3).value = "product"
ws.cell(row = 1, column=4).value = "review"
ws.cell(row = 1, column=5).value = "rating"
ws.cell(row = 1, column=6).value = "price"
wb.save(fn)
r = 2
c = 1

# main program to get the required details
def main():
    r = 2
    c = 1
    for doc in col.find({}):
        wb1 = load_workbook(fn)
        ws = wb1['Sheet1']
        link = doc['link']
        id = doc['_id']
        brand = doc['brand']
        product = doc['product']
        driver.get(link)
        time.sleep(2)
        price = (driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div[2]/div/div[1]/div[2]/div[2 or 3]/div/div[3 or 4]/div[1]/div/div[1]').text)
        price = price.replace('Price: Not Available','0')
        price = price.replace('₹','')
        price = price.replace(',','')
        # print('Price: '+price)
        try:
            rating = (driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/div/div/span[2]/span/span[1]').text)
            rating = rating.replace(' Ratings ','')
            rating = rating.replace(',','')
        except Exception as e:
            rating = '0'
            print(e)
        # print('Rating: ' +rating)
        try:
            review = (driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/div/div/span[2]/span/span[3]').text)
            review = review.replace(' Reviews','')
            review = review.replace(' ','')
            review = review.replace(',','')
        except Exception as e:
            review = '0'
            print(e)
        # print('Reviews: '+review)

        # setting up values to use it as objects
        # rev = str(dt)+'review'
        # rat = str(dt)+'rating'
        # pri = str(dt)+'price'
        # some = {rev:review,rat:rating,pri:price}
        # col.update_one({'brand':brand,'product':product},{'$set': {rev : review, rat : rating, pri : price}})
        # col.update_one({'brand':brand,'product':product},{pri:int(price),rat:int(rating),rev:int(review)},upsert = True)
        # col.update_one({'id':id},{'$set':some},upsert=True)

        # writing to a workbook
        ws.cell(row=r, column=c).value = id
        ws.cell(row=r, column=c+1).value = brand
        ws.cell(row=r, column=c+2).value = product
        ws.cell(row=r, column=c+3).value = review
        ws.cell(row=r, column=c+4).value = rating
        ws.cell(row=r, column=c+5).value = price
        wb1.save(fn)
        r = r+1
        print("Updated:"+str(id)+" : "+str(rating)+" "+str(review)+" "+str(price))
        # upload(id,brand,product,price,rating,review)

    driver.quit()
    # read from the spreadsheet and update into the database
    workbook = xlrd.open_workbook(fn)
    worksheet = workbook.sheet_by_name('Sheet1')
    for val in range(1,worksheet.nrows):
        sid = (worksheet.cell(val,0)).value
        sbrand= (worksheet.cell(val,0)).value
        sproduct = (worksheet.cell(val,0)).value
        c_review = (worksheet.cell(val,0)).value
        c_rating = (worksheet.cell(val,0)).value
        c_price = (worksheet.cell(val,0)).value
        for doc in col.find({}):
            id = doc['_id']
            brand = doc['brand']
            product = doc['product']
            rev = str(dt)+'review'
            rat = str(dt)+'rating'
            pri = str(dt)+'price'
            col.update_one({'id':sid,'brand':sbrand,'product':sproduct},{'$set': {rev : c_review, rat : c_rating, pri : c_price}})

    print("Complete")

if __name__ == '__main__':
    main()
